from __future__ import annotations
import csv
import html
import os
import re
import time
import random
from configparser import ConfigParser
from datetime import datetime
from pathlib import Path
from typing import Optional, List
from openpyxl import load_workbook
from selenium import webdriver
from typing import List, Tuple
from selenium.common.exceptions import (
    TimeoutException,
    NoSuchElementException,
    ElementNotInteractableException,
)
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from tkinter import Tk, Toplevel, Label, StringVar, messagebox
from tkinter import ttk as tkttk


# ─── Configuration ─────────────────────────────────────────────
CONFIG_FILE_PATH = Path(r"C:\Users\nsikder\Downloads\config.ini")
LOG_HTML_FILE = Path("validation_log.html")

# ─── Global log store ──────────────────────────────────────────
log_entries: list[str] = []
html_report_written: bool = False


# ─── Logging Utilities ─────────────────────────────────────────
def log(message: str) -> None:
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    entry = f"[{ts}] {message}"
    log_entries.append(entry)
    print(entry)


# ─── HTML Report ───────────────────────────────────────────────
def save_logs_to_html(
    customer: str,
    export_type: str,
    filename: Path = LOG_HTML_FILE,
    sample_table_html: Optional[str] = None,
) -> None:
    global html_report_written

    try:
        with filename.open("w", encoding="utf-8") as f:
            f.write("<!doctype html><html><head><meta charset='utf-8'>")
            f.write("<title>Export Validation Log</title>")
            f.write("""
            <style>
                body{font-family:Segoe UI,Arial;background:#f6f7f9;padding:20px}
                .card{background:#fff;padding:20px;border-radius:8px}
                .entry{background:#f3f4f6;padding:8px;margin:6px 0;border-radius:6px}
                h1{color:#2f6f17}
            </style>
            </head><body>
            """)

            f.write("<div class='card'>")
            f.write("<h1>Export Dashboard Validation Log</h1>")
            f.write(f"<p><b>Customer:</b> {html.escape(customer)}</p>")
            f.write(f"<p><b>Export:</b> {html.escape(export_type)}</p>")
            f.write("<hr>")

            for e in log_entries:
                f.write(f"<div class='entry'>{html.escape(e)}</div>")

            if sample_table_html:
                f.write(sample_table_html)

            f.write("</div></body></html>")

        html_report_written = True
        log(f"HTML report saved: {filename.resolve()}")

    except Exception as e:
        log(f"Failed to save HTML log: {e}")


# ─── Config & Driver Setup ─────────────────────────────────────
class ConfParser:
    def __init__(self, config_file_path: Path) -> None:
        # 🔧 Disable interpolation to allow % in URLs
        self.config = ConfigParser(interpolation=None)

        if not config_file_path.exists():
            raise FileNotFoundError(f"Config not found: {config_file_path}")

        self.config.read(config_file_path)
        log(f"Config loaded: {config_file_path}")


class ChromeDriverSetup(ConfParser):
    def __init__(self, config_file_path: Path):
        super().__init__(config_file_path)

        options = webdriver.ChromeOptions()

        try:
            options.add_argument(self.config["path"]["chrome_profile"])
        except Exception:
            pass

        prefs = {"safebrowsing.enabled": True}
        options.add_experimental_option("prefs", prefs)

        service = Service(self.config["path"]["chrome_driver"])
        self.driver = webdriver.Chrome(service=service, options=options)

        log("Chrome driver initialized")


def get_usernames_for_customer(customer_name: str) -> Tuple[List[str], str | None]:
    """
    Reads CustomerDB.xlsx
    Column A -> Customer Name
    Column D -> Username
    Returns:
        - list of usernames
        - first username found (or None)
    """
    excel_path = Path("CustomerDB.xlsx")

    if not excel_path.exists():
        raise FileNotFoundError("CustomerDB.xlsx not found in code directory")

    wb = load_workbook(excel_path)
    sheet = wb.active

    customer_norm = normalize_text(customer_name)
    usernames: List[str] = []
    first_username: str | None = None

    for row in sheet.iter_rows(min_row=2):  # skip header
        col_a = row[0].value   # Column A
        col_d = row[3].value   # Column D

        if not col_a or not col_d:
            continue

        if normalize_text(str(col_a)) == customer_norm:
            username = str(col_d).strip()
            usernames.append(username)

            if first_username is None:
                first_username = username  # 👈 store first hit

    return usernames, first_username


# ─── Helper Functions ──────────────────────────────────────────
class SupportiveFunctions:
    driver: webdriver.Chrome
    config: ConfigParser

    def ajax_preloader_wait(self, timeout: int = 300) -> None:
        try:
            time.sleep(1)
            WebDriverWait(self.driver, timeout).until(
                EC.invisibility_of_element(
                    (By.XPATH, "//div[contains(@class,'ajax_preloader')]")
                )
            )
        except Exception:
            log("Ajax preloader wait skipped or timed out")

    def get_element_from_config(self, section: str, key: str):
        xpath = self.config.get(section, key)
        return self.driver.find_element(By.XPATH, xpath)

    def click_from_config(self, section: str, key: str, timeout: int = 20) -> None:
        xpath = self.config.get(section, key)
        WebDriverWait(self.driver, timeout).until(
            EC.element_to_be_clickable((By.XPATH, xpath))
        ).click()

    def send_keys_from_config(
        self, section: str, key: str, value: str, timeout: int = 20
    ) -> None:
        xpath = self.config.get(section, key)
        el = WebDriverWait(self.driver, timeout).until(
            EC.presence_of_element_located((By.XPATH, xpath))
        )
        el.clear()
        el.send_keys(value)



# ─── Progress Window ───────────────────────────────────────────
class ProgressWindow:
    def __init__(self, master: Tk, total_steps: int) -> None:
        self.window = Toplevel(master)
        self.window.title("Validation Progress")
        self.window.geometry("700x180")
        self.window.configure(bg="#4f8611")
        self.window.resizable(False, False)

        self.total_steps = max(1, total_steps)
        self.step = 0

        self.status_var = StringVar(value="Starting...")
        Label(
            self.window,
            textvariable=self.status_var,
            bg="#4f8611",
            fg="white",
            font=("Arial", 13, "bold"),
            wraplength=640,
            justify="center",
        ).pack(pady=10)

        self.progress = tkttk.Progressbar(
            self.window,
            orient="horizontal",
            length=560,
            mode="determinate",
            maximum=self.total_steps,
        )
        self.progress.pack(pady=6)

        self.percent = Label(
            self.window,
            text="0%",
            bg="#4f8611",
            fg="white",
            font=("Arial", 11, "bold"),
        )
        self.percent.pack()

        self.window.update()

    def update(self, message: str) -> None:
        self.step = min(self.step + 1, self.total_steps)
        self.status_var.set(message)
        self.progress["value"] = self.step
        self.percent.config(text=f"{int((self.step / self.total_steps) * 100)}%")
        log(message)
        self.window.update_idletasks()
        time.sleep(0.3)

    def complete(self) -> None:
        self.status_var.set("✅ Validation completed")
        self.progress["value"] = self.total_steps
        self.percent.config(text="100%")
        log("Validation completed")
        self.window.update_idletasks()
        time.sleep(1)
        self.window.destroy()

    def normalize_text(value: str) -> str:
        return " ".join(value.strip().lower().split())


# ─── Login + Validation Runner ─────────────────────────────────
class CozevaLogin(ChromeDriverSetup, SupportiveFunctions):
    def certlogin_cozeva(self, customer: str, progress: ProgressWindow) -> None:
        """Perform login to CERT and select customer via UI interactions."""
        try:
            progress.update("Logging into Cozeva (CERT)...")
            self.driver.get(self.config.get("cert", "logout_url", fallback="about:blank"))
            self.driver.get(self.config.get("cert", "login_url", fallback="about:blank"))
            self.driver.maximize_window()

            user = os.environ.get("CS2User")
            pwd = os.environ.get("CS2Password")
            if not all((user, pwd)):
                raise RuntimeError("Environment variables CS2User / CS2Password not set.")

            self.driver.find_element(By.ID, "edit-name").send_keys(user)
            self.driver.find_element(By.ID, "edit-pass").send_keys(pwd)
            self.driver.find_element(By.ID, "edit-submit").click()

            WebDriverWait(self.driver, 120).until(EC.presence_of_element_located((By.ID, "reason_textbox")))
            self.driver.find_element(By.XPATH, "//*[@id='select-customer']").click()
            self.driver.find_element(By.XPATH, f"//*[contains(text(), '{str(customer)}')]").click()

            WebDriverWait(self.driver, 60).until(EC.presence_of_element_located((By.ID, "reason_textbox")))
            reason_text = self.config.get("credentials", "user_search_reason", fallback="")
            self.driver.find_element(By.ID, "reason_textbox").send_keys(reason_text)
            self.driver.find_element(By.ID, "edit-submit").click()
            time.sleep(5)

            self.ajax_preloader_wait()
            progress.update("Logged in successfully (CERT).")
        except Exception as e:
            log(f"❌ Login Error (CERT): {e}")
            messagebox.showerror("Login Error", str(e))
            raise

    def prodlogin_cozeva(self, customer: str, progress: ProgressWindow) -> None:
        """Perform login to PROD and select customer via UI interactions."""
        try:
            progress.update("Logging into Cozeva (PROD)...")
            self.driver.get(self.config.get("prod", "logout_url", fallback="about:blank"))
            self.driver.get(self.config.get("prod", "login_url", fallback="about:blank"))
            self.driver.maximize_window()

            user = os.environ.get("CS2User")
            pwd = os.environ.get("CS2Password")
            if not all((user, pwd)):
                raise RuntimeError("Environment variables CS2User / CS2Password not set.")

            self.driver.find_element(By.ID, "edit-name").send_keys(user)
            self.driver.find_element(By.ID, "edit-pass").send_keys(pwd)
            self.driver.find_element(By.ID, "edit-submit").click()

            WebDriverWait(self.driver, 120).until(EC.presence_of_element_located((By.ID, "reason_textbox")))
            self.driver.find_element(By.XPATH, "//*[@id='select-customer']").click()
            self.driver.find_element(By.XPATH, f"//*[contains(text(), '{str(customer)}')]").click()

            WebDriverWait(self.driver, 60).until(EC.presence_of_element_located((By.ID, "reason_textbox")))
            reason_text = self.config.get("credentials", "user_search_reason", fallback="")
            self.driver.find_element(By.ID, "reason_textbox").send_keys(reason_text)
            self.driver.find_element(By.ID, "edit-submit").click()
            time.sleep(5)

            self.ajax_preloader_wait()
            progress.update("Logged in successfully (PROD).")
        except Exception as e:
            log(f"❌ Login Error (PROD): {e}")
            messagebox.showerror("Login Error", str(e))
            raise

    def logout(self, progress: ProgressWindow, customer: str) -> None:
        try:
            self.driver.quit()
        except Exception:
            pass
        progress.complete()
        save_logs_to_html(customer, "User Search Validation")


def normalize_text(value: str) -> str:
    return " ".join(value.strip().lower().split())


class user_search(SupportiveFunctions):
    def __init__(self, driver: webdriver.Chrome, config: ConfigParser):
        self.driver = driver
        self.config = config

    def users_list(self, customer: str, progress: ProgressWindow, usernames: List[str]) -> None:
        try:
            progress.update("Opening User List page...")

            # 1️⃣ Open User List page
            self.driver.get(self.config.get("user_list", "list_url", fallback="about:blank"))
            self.ajax_preloader_wait()
            time.sleep(2)
            progress.update("Opening user list filter...")
            self.click_from_config("UserListLocator", "xpath_userlist_filter")
            time.sleep(0.5)

            # 2️⃣ Open customer dropdown
            progress.update("Opening customer dropdown...")
            self.click_from_config("UserListLocator", "xpath_customername")
            time.sleep(0.5)

            # 3️⃣ Select customer
            customer_norm = normalize_text(customer)
            log(f"Normalized customer text: '{customer_norm}'")

            dropdown_ul = WebDriverWait(self.driver, 20).until(EC.visibility_of_element_located((By.XPATH, "//ul[contains(@class,'select-dropdown') and contains(@style,'display')]")))

            matched_element = None
            for opt in dropdown_ul.find_elements(By.TAG_NAME, "li"):
                raw_text = opt.text.strip()
                if raw_text and normalize_text(raw_text) == customer_norm:
                    matched_element = opt
                    log(f"Matched dropdown option: {raw_text}")
                    break

            if not matched_element:
                raise ValueError(f"Customer '{customer}' not found in dropdown")
            matched_element.click()
            progress.update(f"Customer '{customer}' selected successfully.")


            # 5️⃣ Username loop
            for idx, username in enumerate(usernames):
                log(f"Applying username filter: {username}")

                # Re-open filter for subsequent usernames
                if idx > 0:
                    progress.update("Re-opening user list filter...")
                    self.click_from_config("UserListLocator", "xpath_userlist_filter")
                    time.sleep(0.5)

                search_input = WebDriverWait(self.driver, 20).until(
                    EC.element_to_be_clickable((By.XPATH, "//input[@name='search_people']")))

                search_input.clear()
                search_input.send_keys(username)

                WebDriverWait(self.driver, 20).until(EC.element_to_be_clickable((By.XPATH, "//a[contains(@class,'datatable_apply')]"))).click()
                self.ajax_preloader_wait()

                try:
                    result_cell = WebDriverWait(self.driver, 20).until(
                        EC.presence_of_element_located(
                            (By.XPATH, "//td[@class='username username_pt sorting_1']")
                        )
                    )
                except TimeoutException:
                    log(f"❌ No result row found for username '{username}'")
                    continue

                ui_username = result_cell.text.strip()

                if normalize_text(ui_username) == normalize_text(username):
                    log(f"✅ Username matched: UI='{ui_username}' | Expected='{username}'")
                else:
                    log(f"❌ Username mismatch: UI='{ui_username}' | Expected='{username}'")

        except Exception as e:
            log(f"❌ Failed to open User List or select customer: {e}")
            raise

    def batch_share(self, customer: str, progress: ProgressWindow, first_username: str | None) -> None:
        try:
            if not first_username:
                raise ValueError("First username is None. Cannot search batch.")
            progress.update("Opening Batch List page...")

            # 1️⃣ Open Batch List page
            self.driver.get(self.config.get("batch_list", "batch_url", fallback="about:blank"))
            self.ajax_preloader_wait()
            time.sleep(2)
            self.click_from_config("BatchListLocator", "xpath_batch_menu")
            time.sleep(2)
            self.click_from_config("BatchListLocator", "xpath_batch_share")
            time.sleep(2)
            # 2️⃣ Search using first_username
            search_field = self.get_element_from_config("BatchListLocator", "xpath_batch_search")
            search_field.clear()
            search_field.send_keys(first_username)
            username_elem = WebDriverWait(self.driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, "//ul[@id='ac-dropdown-share-with']//li[1]//b")))
            ui_username = username_elem.text.strip()
            print("UI Username:", ui_username)
            progress.update(ui_username)

            # Compare with first_username
            if ui_username == first_username:
                print("✅ Username matches")
            else:
                print("❌ Username mismatch")
            time.sleep(5)

        except Exception as e:
            log(f"❌ Failed to open Batch List or select customer: {e}")
            raise

    def secure_messaging(self, customer: str, progress: ProgressWindow, first_username: str | None) -> None:
        try:
            if not first_username:
                raise ValueError("First username is None. Cannot search batch.")
            progress.update("Opening Batch List page...")

            # 1️⃣ Open Batch List page
            self.driver.get(self.config.get("secure_messaging", "secure_url", fallback="about:blank"))
            self.ajax_preloader_wait()
            time.sleep(2)
            self.click_from_config("SecureMessagingLocator", "xpath_new_message")
            time.sleep(2)
            self.click_from_config("SecureMessagingLocator", "xpath_select_dropdown")
            time.sleep(2)
            self.click_from_config("SecureMessagingLocator", "xpath_customer_support")
            time.sleep(2)
            secrch_username = WebDriverWait(self.driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, "//input[@data-drupal-selector='edit-proname']")))
            secrch_username.click()
            secrch_username.send_keys(first_username)
            username_elem = WebDriverWait(self. driver, 10).until(EC.visibility_of_element_located((By.XPATH, "//ul[@id='ac-dropdown-share-with']//li[1]//b")))
            ui_username = username_elem.text.strip()
            print("UI Username:", ui_username)

            # Compare with first_username
            if ui_username == first_username:
                print("✅ Username matches")
            else:
                print("❌ Username mismatch")
            time.sleep(5)

        except Exception as e:
            log(f"❌ Failed to open Batch List or select customer: {e}")
            raise

    def analytics_search(self, customer: str, progress: ProgressWindow, first_username: str | None) -> None:
        try:
            if not first_username:
                raise ValueError("First username is None. Cannot search batch.")
            progress.update("Opening Analytics...")

            # 1️⃣ Open Analytics
            self.driver.get(self.config.get("analytics", "analytics_url", fallback="about:blank"))
            self.ajax_preloader_wait()
            time.sleep(2)
            self.click_from_config("AnalyticsLocator", "xpath_analytics_share")
            time.sleep(5)
            self.click_from_config("AnalyticsLocator", "xpath_analytics_dropdown")
            time.sleep(2)
            # 2️⃣ Search using first_username
            search_field1 = self.get_element_from_config("AnalyticsLocator", "xpath_user_search")
            search_field1.clear()
            search_field1.send_keys(first_username)
            first_result = WebDriverWait(self.driver, 10).until(
                EC.visibility_of_element_located(
                    (
                        By.XPATH,
                        "(//ul[contains(@class,'multiselect-container')])[24]//li[contains(@class,'context1') and contains(@style,'display: block')]"
                    )
                )
            )

            raw_text = first_result.text.strip()
            match = re.search(r"\(([^)]+)\)", raw_text)
            ui_username1 = match.group(1).strip() if match else ""
            print("UI Username (first result):", ui_username1)
            progress.update(ui_username1)

            # Compare with first_username
            if ui_username1 == first_username:
                print("✅ Username matches")
            else:
                print("❌ Username mismatch")
            time.sleep(5)

        except Exception as e:
            log(f"❌ Failed to open Analytics or select customer: {e}")
            raise

    def ticket_search(self, customer: str, progress: ProgressWindow) -> None:
        progress.update("Opening Support Ticket Page...")
        self.driver.get(self.config.get("support_ticket", "ticket_url", fallback="about:blank"))
        self.ajax_preloader_wait()
        time.sleep(2)
        plus_xpath = "//a[@class='btn-floating btn-large red waves-effect waves-light new_support_activity_btn']"
        WebDriverWait(self.driver, 60).until(EC.visibility_of_element_located((By.XPATH, plus_xpath)))
        self. driver.find_element(By.XPATH, plus_xpath).click()
        time.sleep(2)
        self.ajax_preloader_wait()
        self.driver.find_element(By.XPATH, "(//i[@class='tiny material-icons ac-icon ac-clear'])[2]").click()
        time.sleep(3)
        self.driver.find_element(By.XPATH, '(//input[@name="assignee"])').send_keys("Aritra")
        time.sleep(5)
        first_item = WebDriverWait(self.driver, 10).until(
            EC.element_to_be_clickable(
                (
                    By.XPATH,
                    "(//ul[@class='dropdown-content mat-ac-dropdown ']//li[@tabindex='0'])[1]"
                )
            )
        )
        output_text = first_item.text.strip()
        print("Dropdown first item:", output_text)
        progress.update(output_text)

        expected_text = "Aritra Mukherjee | Cozeva Support | amukherjee.cs"
        if output_text == expected_text:
            print("✅ Text matches exactly")
        else:
            print(f"❌ Mismatch\nExpected: {expected_text}\nFound: {output_text}")

    def casemanagement_search(self, customer: str, progress: ProgressWindow):
        progress.update("Opening patient dashboard to perform Case Management User Search...")
        # 1️⃣ Open Batch List page
        self.driver.get(self.config.get("case_management", "task_url", fallback="about:blank"))
        self.ajax_preloader_wait()
        time.sleep(2)
        self.click_from_config("CMLocator", "xpath_kebab_icon")
        time.sleep(2)
        self.click_from_config("CMLocator", "xpath_edit_task")
        time.sleep(5)
        self.ajax_preloader_wait()
        user_name = "avijit CozevaQA"
        search_field2 = self.get_element_from_config("CMLocator", "xpath_cm_assignee")
        time.sleep(2)
        search_field2.send_keys(user_name)
        username_elem1 = WebDriverWait(self.driver, 10).until(
            EC.visibility_of_element_located((By.XPATH, "//ul[@id='ac-dropdown-edit-edit-assignee-name']")))

        ui_username1 = username_elem1.text.split("(", 1)[0].strip()
        print("UI Username:", ui_username1)
        progress.update(ui_username1)

        # Compare with first_username
        if ui_username1 == user_name:
            print("✅ Username matches")
        else:
            print("❌ Username mismatch")
        time.sleep(5)

    def deletetestingdata_search(self, customer: str, progress: ProgressWindow, usernames: List[str]) -> None:
        try:
            progress.update("Opening Support Tool list...")

            # 1️⃣ Open Batch List page
            self.driver.get(self.config.get("delete_data", "supporttool_url", fallback="about:blank"))
            self.ajax_preloader_wait()
            time.sleep(2)
            self.click_from_config("SupportToolLocator", "xpath_deletetest_data")
            time.sleep(5)
            self.click_from_config("SupportToolLocator", "xpath_masq_checkbox")
            time.sleep(2)
            for idx, username in enumerate(usernames):
                log(f"Searching username: {username}")
                # Get search field
                search_field = self.get_element_from_config("SupportToolLocator", "xpath_deletedata_user")
                search_field.clear()
                time.sleep(0.5)
                search_field.send_keys(username)
                time.sleep(2)
                username_elem1 = WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located((By.XPATH, "//ul[@id='ac-dropdown-logged_or_masquaraded_user_name']//li")))
                # Extract UI username (before | )
                ui_username1 = username_elem1.text.split("|", 1)[0].strip()

                print("UI Username:", ui_username1)
                progress.update(ui_username1)

                # Compare
                if normalize_text(ui_username1) == normalize_text(username):
                    log(f"✅ Username matched: UI='{ui_username1}' | Expected='{username}'")
                else:
                    log(f"❌ Username mismatch: UI='{ui_username1}' | Expected='{username}'")

                time.sleep(2)
        except Exception as e:
            log(f"❌ Failed to open Batch List or select customer: {e}")
            raise


# ─── ENTRY POINT CALLED FROM UI ─────────────────────────────────
def run_user_validation(
    master_window: Tk,
    customer: str,
    selected_areas: List[str],
    environment: str = "CERT",
) -> None:

    # 🔒 Safety: ensure list
    if not isinstance(selected_areas, list):
        selected_areas = [selected_areas]

    progress = ProgressWindow(master_window, total_steps=3 + len(selected_areas))

    try:
        # ─── Driver + Login ────────────────────────────────────
        runner = CozevaLogin(CONFIG_FILE_PATH)

        if environment.upper() == "CERT":
            runner.certlogin_cozeva(customer, progress)
        else:
            raise NotImplementedError("PROD login not wired yet")

        # ─── Initialize Search Handler ─────────────────────────
        search = user_search(runner.driver, runner.config)

        # ─── Fetch usernames ───────────────────────────────────
        usernames, first_username = get_usernames_for_customer(customer)

        if not usernames:
            raise ValueError(f"No usernames found for '{customer}' in CustomerDB.xlsx")

        # ─── Ordered execution map ─────────────────────────────
        execution_flow = [
            ("User List", lambda: search.users_list(customer, progress, usernames)),
            ("Batch Share", lambda: search.batch_share(customer, progress, first_username)),
            ("Analytics", lambda: search.analytics_search(customer, progress, first_username)),
            ("Support Ticket", lambda: search.ticket_search(customer, progress)),
            ("Case Management", lambda: search.casemanagement_search(customer, progress)),
            ("Delete Testing Data", lambda: search.deletetestingdata_search(customer, progress, usernames)),
        ]

        executed_any = False

        # ─── Execute selected areas IN ORDER ───────────────────
        for area_name, action in execution_flow:
            if area_name in selected_areas:
                executed_any = True
                progress.update(f"Starting User Search validation in {area_name}...")
                action()

        # ─── Nothing selected ──────────────────────────────────
        if not executed_any:
            log("ℹ️ User Search skipped (no valid area selected)")
            messagebox.showinfo(
                "Selection Required",
                "Please select at least one validation area."
            )

        # ─── Logout ────────────────────────────────────────────
        runner.logout(progress, customer)

    except Exception as e:
        log(f"❌ Validation failed: {e}")
        messagebox.showerror("Validation Error", str(e))
        try:
            progress.window.destroy()
        except Exception:
            pass

